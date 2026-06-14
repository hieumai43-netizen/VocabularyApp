import openpyxl
import json

excel_file = "vocabulary_database_template.xlsx"
output_file = "data.js"
sheet_name = "Vocabulary"

wb = openpyxl.load_workbook(excel_file)
ws = wb[sheet_name]

headers = {}
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    if header:
        headers[str(header).strip()] = col

def get_value(row, column_name):
    col = headers.get(column_name)
    if not col:
        return ""
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()

data = []

for row in range(2, ws.max_row + 1):
    japanese = get_value(row, "Japanese")
    english = get_value(row, "English")
    chinese = get_value(row, "Chinese")
    korean = get_value(row, "Korean")
    vietnamese = get_value(row, "Vietnamese")

    if not japanese and not english and not chinese and not korean and not vietnamese:
        continue

    item = {
        "id": get_value(row, "ID"),
        "category": get_value(row, "Category"),
        "vi": vietnamese,
        "jp": japanese,
        "kana": get_value(row, "Kana"),
        "en": english,
        "ipa": get_value(row, "IPA"),
        "cn": chinese,
        "pinyin": get_value(row, "Pinyin"),
        "ko": korean,
        "koreanReading": get_value(row, "KoreanReading"),
        "note": get_value(row, "Note"),
        "examples": {
            "vi": get_value(row, "Example_VI"),
            "jp": get_value(row, "Example_JA"),
            "en": get_value(row, "Example_EN"),
            "cn": get_value(row, "Example_CN"),
            "ko": get_value(row, "Example_KO")
        }
    }

    data.append(item)

js_content = "const vocabularyData = "
js_content += json.dumps(data, ensure_ascii=False, indent=2)
js_content += ";"

with open(output_file, "w", encoding="utf-8") as f:
    f.write(js_content)

print("Đã tạo xong file data.js")
print(f"Tổng số từ: {len(data)}")